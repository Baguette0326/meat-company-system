from __future__ import annotations

import tempfile
import time
import shutil
from datetime import datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook, load_workbook
from pypdf import PdfReader
from master_data import refresh_order_template

from read_daily_orders import (
    detect_duplicate_orders,
    generate_daily_report,
    generate_monthly_report,
    read_order_file,
)


def make_order(
    path: Path,
    order_date: str,
    customer: str,
    order_number: str,
    entries: list[tuple[int, str, str, Decimal, str, Decimal]],
    status: str = "正常",
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "訂單輸入"
    sheet["B5"] = customer
    sheet["C5"] = status
    sheet["F5"] = datetime.strptime(order_date, "%Y%m%d")
    sheet["H5"] = order_number
    for row, product, category, quantity, price_column, price in entries:
        sheet.cell(row, 1, float(quantity))
        sheet.cell(row, 3, product)
        sheet.cell(row, 4, category)
        sheet.cell(row, 5, product)
        sheet[f"{price_column}{row}"] = float(price)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def find_summary_value(path: Path, label: str) -> Decimal:
    workbook = load_workbook(path, data_only=True)
    sheet = workbook["總覽"]
    for row in sheet.iter_rows():
        if row[0].value == label:
            return Decimal(str(row[1].value))
    raise AssertionError(f"Missing summary label: {label}")


def require(condition: bool, message: str) -> None:
    if not condition:
        raise AssertionError(message)


def run() -> None:
    started = time.perf_counter()
    with tempfile.TemporaryDirectory(prefix="sales-preflight-") as temp:
        root = Path(temp)
        orders = root / "Orders"
        current = orders / "20260618"
        previous = orders / "20260518"

        current_file = current / "20260618_測試客戶_001.xlsx"
        make_order(
            current_file,
            "20260618",
            "測試客戶",
            "001",
            [
                (9, "封門柳/薄肉/頸皮", "牛", Decimal("2"), "F", Decimal("10")),
                (53, "豬腰", "豬", Decimal("3"), "G", Decimal("20")),
                (84, "羊肉", "羊", Decimal("4"), "H", Decimal("30")),
                (89, "尾龍骨", "雜貨", Decimal("5"), "F", Decimal("4")),
            ],
        )
        make_order(
            previous / "20260518_測試客戶_001.xlsx",
            "20260518",
            "測試客戶",
            "001",
            [(9, "封門柳/薄肉/頸皮", "牛", Decimal("1"), "F", Decimal("10"))],
        )

        rows, issues = read_order_file(current_file, datetime(2026, 6, 18).date())
        require(not issues, f"Valid order produced issues: {issues}")
        require(sum((row.revenue for row in rows), Decimal("0")) == Decimal("220.00"), "F/G/H revenue calculation failed")
        require({row.price_type for row in rows} == {"平均售價", "平價切", "精修切"}, "Price-type detection failed")

        wrong_date_rows, wrong_date_issues = read_order_file(current_file, datetime(2026, 6, 19).date())
        require(wrong_date_rows == rows, "Date warning changed valid sales data")
        require(any("資料夾日期" in issue.message for issue in wrong_date_issues), "Wrong-folder-date warning missing")

        duplicate_file = current / "20260618_測試客戶_002.xlsx"
        make_order(
            duplicate_file,
            "20260618",
            "測試客戶",
            "001",
            [(9, "封門柳/薄肉/頸皮", "牛", Decimal("2"), "F", Decimal("10"))],
        )
        duplicate_rows, _ = read_order_file(duplicate_file)
        duplicate_issues = detect_duplicate_orders(rows + duplicate_rows)
        require(any("重複訂單編號" in issue.message for issue in duplicate_issues), "Duplicate-order warning missing")
        duplicate_file.unlink()

        canceled_file = current / "20260618_取消測試_003.xlsx"
        make_order(
            canceled_file,
            "20260618",
            "取消測試",
            "003",
            [(9, "封門柳/薄肉/頸皮", "牛", Decimal("99"), "F", Decimal("99"))],
            status="取消",
        )
        canceled_rows, canceled_issues = read_order_file(canceled_file)
        require(canceled_rows == [], "Canceled order should not produce sales rows")
        require(any("不計入銷售" in issue.message for issue in canceled_issues), "Canceled-order notice missing")

        return_file = current / "20260618_退貨測試_004.xlsx"
        make_order(
            return_file,
            "20260618",
            "退貨測試",
            "004",
            [(9, "封門柳/薄肉/頸皮", "牛", Decimal("1"), "F", Decimal("10"))],
            status="退貨",
        )
        return_rows, return_issues = read_order_file(return_file)
        require(not return_issues, f"Return order produced issues: {return_issues}")
        require(sum((row.revenue for row in return_rows), Decimal("0")) == Decimal("-10.00"), "Return revenue should be negative")

        corrected_file = current / "20260618_更正測試_005.xlsx"
        make_order(
            corrected_file,
            "20260618",
            "更正測試",
            "005",
            [(9, "封門柳/薄肉/頸皮", "牛", Decimal("1"), "F", Decimal("5"))],
            status="更正",
        )
        corrected_rows, corrected_issues = read_order_file(corrected_file)
        require(not corrected_issues, f"Corrected order produced issues: {corrected_issues}")
        require(sum((row.revenue for row in corrected_rows), Decimal("0")) == Decimal("5.00"), "Corrected revenue incorrect")

        daily_xlsx = root / "Daily Reports" / "每日報表_20260618.xlsx"
        daily_counts = generate_daily_report(current, daily_xlsx)
        require(daily_counts[0:2] == (4, 6), f"Unexpected daily file/row counts: {daily_counts}")
        require(daily_counts[2] >= 1, f"Canceled-order issue count missing: {daily_counts}")
        require(daily_xlsx.exists() and daily_xlsx.with_suffix(".pdf").exists(), "Daily Excel/PDF missing")
        require(find_summary_value(daily_xlsx, "總收入 HKD") == Decimal("215"), "Daily total revenue incorrect")
        require(find_summary_value(daily_xlsx, "取消訂單數量") == Decimal("1"), "Daily canceled-order summary missing")
        require(find_summary_value(daily_xlsx, "更正訂單數量") == Decimal("1"), "Daily corrected-order summary missing")
        require(find_summary_value(daily_xlsx, "退貨訂單數量") == Decimal("1"), "Daily return-order summary missing")
        daily_issues_sheet = load_workbook(daily_xlsx, data_only=True)["問題"]
        daily_issue_text = "\n".join(str(row[2]) for row in daily_issues_sheet.iter_rows(min_row=2, values_only=True))
        require("訂單狀態為更正" in daily_issue_text, "Daily corrected-order issue notice missing")
        require("訂單狀態為退貨" in daily_issue_text, "Daily return-order issue notice missing")
        daily_backups = list((root / "Backups").glob("daily_20260618_*/reports/每日報表_20260618.xlsx"))
        require(daily_backups, "Daily report backup missing")

        monthly_xlsx = root / "Monthly Reports" / "月報表_202606.xlsx"
        monthly_counts = generate_monthly_report(orders, "202606", monthly_xlsx)
        require(monthly_counts[0:2] == (4, 6), f"Unexpected monthly file/row counts: {monthly_counts}")
        require(monthly_counts[2] >= 1, f"Canceled-order issue count missing in monthly counts: {monthly_counts}")
        require(monthly_xlsx.exists() and monthly_xlsx.with_suffix(".pdf").exists(), "Monthly Excel/PDF missing")
        require(find_summary_value(monthly_xlsx, "總收入 HKD") == Decimal("215"), "Monthly total revenue incorrect")
        require(find_summary_value(monthly_xlsx, "取消訂單數量") == Decimal("1"), "Monthly canceled-order summary missing")
        require(find_summary_value(monthly_xlsx, "更正訂單數量") == Decimal("1"), "Monthly corrected-order summary missing")
        require(find_summary_value(monthly_xlsx, "退貨訂單數量") == Decimal("1"), "Monthly return-order summary missing")
        monthly_issues_sheet = load_workbook(monthly_xlsx, data_only=True)["問題"]
        monthly_issue_text = "\n".join(str(row[2]) for row in monthly_issues_sheet.iter_rows(min_row=2, values_only=True))
        require("訂單狀態為更正" in monthly_issue_text, "Monthly corrected-order issue notice missing")
        require("訂單狀態為退貨" in monthly_issue_text, "Monthly return-order issue notice missing")
        monthly_backups = list((root / "Backups").glob("monthly_202606_*/reports/月報表_202606.xlsx"))
        require(monthly_backups, "Monthly report backup missing")
        comparison = load_workbook(monthly_xlsx, data_only=True)["月度比較"]
        revenue_comparison = next(
            row for row in comparison.iter_rows(min_row=2, values_only=True) if row[0] == "總收入 HKD"
        )
        require(Decimal(str(revenue_comparison[1])) == Decimal("215"), "Current-month comparison revenue incorrect")
        require(Decimal(str(revenue_comparison[2])) == Decimal("10"), "Previous-month comparison revenue incorrect")
        require(len(PdfReader(monthly_xlsx.with_suffix(".pdf")).pages) == 2, "Monthly PDF is not two pages")

        dynamic_template = root / "Order Template" / "order-file-template.xlsx"
        dynamic_master = root / "Master Data" / "master-data.xlsx"
        dynamic_template.parent.mkdir(parents=True, exist_ok=True)
        dynamic_master.parent.mkdir(parents=True, exist_ok=True)
        shutil.copyfile(Path(__file__).resolve().parents[1] / "Order Template" / "order-file-template.xlsx", dynamic_template)

        master_workbook = Workbook()
        product_sheet = master_workbook.active
        product_sheet.title = "貨品清單"
        product_sheet.append(["貨品", "分類", "部位", "單位", "狀態"])
        product_sheet.append(["新增測試貨品", "牛", "新增測試部位", "KG", "使用中"])
        customer_sheet = master_workbook.create_sheet("客戶清單")
        customer_sheet.append(["客戶名稱", "狀態"])
        customer_sheet.append(["測試客戶", "使用中"])
        master_workbook.save(dynamic_master)

        refresh_order_template(dynamic_template, dynamic_master)
        template_workbook = load_workbook(dynamic_template)
        template_sheet = template_workbook[template_workbook.sheetnames[0]]
        new_product_row = next(
            row
            for row in range(7, template_sheet.max_row + 1)
            if template_sheet.cell(row, 3).value == "新增測試貨品"
        )
        template_sheet["B5"] = "測試客戶"
        template_sheet["C5"] = "正常"
        template_sheet["F5"] = datetime.strptime("20260618", "%Y%m%d")
        template_sheet["H5"] = "999"
        template_sheet.cell(new_product_row, 1, 2)
        template_sheet.cell(new_product_row, 6, 50)
        dynamic_order = root / "Orders" / "20260618" / "20260618_測試客戶_999.xlsx"
        template_workbook.save(dynamic_order)
        dynamic_rows, dynamic_issues = read_order_file(dynamic_order)
        require(not dynamic_issues, f"Dynamic template order produced issues: {dynamic_issues}")
        require(sum((row.revenue for row in dynamic_rows), Decimal("0")) == Decimal("100.00"), "Dynamic product row was not read")

    elapsed = time.perf_counter() - started
    print(f"PASS: calculations, warnings, duplicate detection, Excel/PDF reports ({elapsed:.2f}s)")


if __name__ == "__main__":
    run()
