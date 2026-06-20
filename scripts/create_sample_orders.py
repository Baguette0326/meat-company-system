from __future__ import annotations

from copy import copy
from datetime import date
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = Path(__file__).resolve().parents[1] / "Daily Reports" / "order-file-template-updated.xlsx"
OUTPUT_DIR = ROOT / "Orders" / "20260618"


def clone_template(output_path: Path, customer: str, sequence: str, entries: list[tuple[str, float, float, str]]):
    workbook = load_workbook(TEMPLATE)
    sheet = workbook["訂單輸入"] if "訂單輸入" in workbook.sheetnames else workbook[workbook.sheetnames[0]]

    sheet["B5"] = customer
    sheet["F5"] = date(2026, 6, 18)
    sheet["H5"] = sequence

    product_to_row = {
        str(sheet.cell(row, 3).value).strip(): row
        for row in range(1, sheet.max_row + 1)
        if sheet.cell(row, 3).value
    }

    for product, quantity, price, price_col in entries:
        row = product_to_row[product]
        sheet.cell(row, 1).value = quantity
        if price_col == "G":
            sheet.cell(row, 7).value = price
        elif price_col == "H":
            sheet.cell(row, 8).value = price
        else:
            raise ValueError(f"Unsupported price column: {price_col}")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def main() -> None:
    clone_template(
        OUTPUT_DIR / "20260618_海景酒家_001.xlsx",
        "海景酒家",
        "001",
        [
            ("雪花", 7, 18, "G"),
            ("牛腩/硼砂腩", 3.5, 42, "H"),
            ("豬腩", 5, 28, "G"),
        ],
    )
    clone_template(
        OUTPUT_DIR / "20260618_大發餐廳_001.xlsx",
        "大發餐廳",
        "001",
        [
            ("牛柳", 4, 80, "H"),
            ("排骨", 6, 35, "G"),
        ],
    )
    clone_template(
        OUTPUT_DIR / "20260618_大發餐廳_002.xlsx",
        "大發餐廳",
        "002",
        [
            ("羊肉", 2.5, 55, "H"),
            ("臘腸", 10, 22, "G"),
        ],
    )

    print(f"Created samples in {OUTPUT_DIR}")


if __name__ == "__main__":
    main()
