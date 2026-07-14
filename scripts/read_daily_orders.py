from __future__ import annotations

import argparse
import calendar
import shutil
import re
import unicodedata
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

from master_data import MASTER_DATA_PATH, MasterProduct, ensure_master_data, load_master_customers, load_master_products


PRICE_TYPES = {
    "average": "平均售價",
    "cheap": "平價切",
    "fine": "精修切",
}

FILENAME_RE = re.compile(r"^(?P<date>\d{8})_(?P<customer>.+)_(?P<seq>\d{3})\.xlsx$", re.IGNORECASE)


@dataclass(frozen=True)
class OrderFileMeta:
    path: Path
    filename_date: date | None
    filename_customer: str | None
    sequence: str | None


@dataclass(frozen=True)
class SaleRow:
    source_file: str
    row_number: int
    order_date: date | None
    customer: str
    order_number: str
    order_status: str
    category: str
    product: str
    part: str
    quantity: Decimal
    unit_price: Decimal
    price_type: str
    buy_in_price: Decimal | None
    revenue: Decimal
    estimated_margin: Decimal | None


@dataclass(frozen=True)
class ImportIssue:
    source_file: str
    row_number: int | None
    severity: str
    message: str


COUNTED_ORDER_STATUSES = {"正常", "更正", "退貨"}
ORDER_STATUS_ALIASES = {
    "": "正常",
    "正常": "正常",
    "取消": "取消",
    "已取消": "取消",
    "更正": "更正",
    "退貨": "退貨",
    "退货": "退貨",
}


def to_decimal(value) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value).strip())
    except (InvalidOperation, AttributeError):
        return None


def money(value: Decimal | None) -> Decimal | None:
    if value is None:
        return None
    return value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def parse_filename(path: Path) -> OrderFileMeta:
    match = FILENAME_RE.match(path.name)
    if not match:
        return OrderFileMeta(path, None, None, None)

    raw_date = match.group("date")
    try:
        parsed_date = datetime.strptime(raw_date, "%Y%m%d").date()
    except ValueError:
        parsed_date = None

    return OrderFileMeta(
        path=path,
        filename_date=parsed_date,
        filename_customer=match.group("customer"),
        sequence=match.group("seq"),
    )


def normalize_date(value) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value is None or value == "":
        return None
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def cell_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_order_status(value) -> str:
    return ORDER_STATUS_ALIASES.get(cell_text(value), cell_text(value))


def get_sheet(workbook):
    if "訂單輸入" in workbook.sheetnames:
        return workbook["訂單輸入"]
    if "每日銷售表" in workbook.sheetnames:
        return workbook["每日銷售表"]
    return workbook[workbook.sheetnames[0]]


def read_header_fields(sheet):
    customer = cell_text(sheet["B5"].value)
    order_status = normalize_order_status(sheet["C5"].value)
    order_date = normalize_date(sheet["F5"].value)
    order_number = cell_text(sheet["H5"].value)

    # Some users may enter the values beside the labels on row 5, but keep a
    # fallback search in case the layout is slightly shifted.
    if not customer or not order_status or not order_date or not order_number:
        for row in sheet.iter_rows(min_row=1, max_row=8, min_col=1, max_col=8):
            values = [cell_text(cell.value) for cell in row]
            for index, value in enumerate(values):
                next_value = values[index + 1] if index + 1 < len(values) else ""
                if not customer and value.startswith("客戶名稱"):
                    customer = next_value
                if not order_status and value.startswith("訂單狀態"):
                    order_status = normalize_order_status(next_value)
                if not order_date and value.startswith("日期"):
                    order_date = normalize_date(next_value)
                if not order_number and value.startswith("發貨單編號"):
                    order_number = next_value

    return customer, order_status or "正常", order_date, order_number


def is_section_or_header(row_values: list[str]) -> bool:
    first = row_values[0]
    product = row_values[2]
    if first in {"牛", "豬", "鱼", "魚", "羊", "雜貨", "杂货"}:
        return True
    if first == "售出数量" or product == "品名":
        return True
    if first in {"總結", "訂單總結", "每日總結"}:
        return True
    return False


def iter_workbooks(folder: Path) -> Iterable[Path]:
    for path in sorted(folder.glob("*.xlsx")):
        if path.name.startswith("~$"):
            continue
        yield path


def folder_date(folder: Path) -> date | None:
    if not re.fullmatch(r"\d{8}", folder.name):
        return None
    try:
        return datetime.strptime(folder.name, "%Y%m%d").date()
    except ValueError:
        return None


def read_order_file(
    path: Path,
    expected_folder_date: date | None = None,
    master_products: dict[str, MasterProduct] | None = None,
    master_customers: set[str] | None = None,
) -> tuple[list[SaleRow], list[ImportIssue]]:
    issues: list[ImportIssue] = []
    rows: list[SaleRow] = []
    meta = parse_filename(path)

    if meta.filename_date is None or meta.filename_customer is None:
        issues.append(
            ImportIssue(
                path.name,
                None,
                "warning",
                "檔名應使用 YYYYMMDD_客戶名_序號.xlsx",
            )
        )

    try:
        workbook = load_workbook(path, data_only=False)
    except Exception as exc:  # noqa: BLE001 - report workbook-level import failure.
        return [], [ImportIssue(path.name, None, "錯誤", f"無法開啟 Excel 檔案: {exc}")]

    sheet = get_sheet(workbook)
    customer, order_status, order_date, order_number = read_header_fields(sheet)

    if not customer:
        issues.append(ImportIssue(path.name, None, "錯誤", "B5 缺少客戶名稱"))
    elif master_customers is not None and master_customers and customer not in master_customers:
        issues.append(ImportIssue(path.name, None, "警告", f"客戶「{customer}」不在客戶清單內"))
    if order_status not in ORDER_STATUS_ALIASES.values():
        issues.append(
            ImportIssue(
                path.name,
                None,
                "錯誤",
                f"C5 訂單狀態「{order_status}」無效，請使用：正常、取消、更正、退貨",
            )
        )
    if not order_date:
        issues.append(ImportIssue(path.name, None, "錯誤", "F5 缺少日期或日期格式無效"))
    if not order_number:
        order_number = meta.sequence or ""
        issues.append(ImportIssue(path.name, None, "警告", "H5 缺少發貨單編號"))

    if meta.filename_date and order_date and meta.filename_date != order_date:
        issues.append(
            ImportIssue(
                path.name,
                None,
                "警告",
                f"檔名日期 {meta.filename_date.isoformat()} 與 Excel 日期 {order_date.isoformat()} 不一致",
            )
        )
    if expected_folder_date and meta.filename_date and meta.filename_date != expected_folder_date:
        issues.append(
            ImportIssue(
                path.name,
                None,
                "警告",
                f"檔名日期 {meta.filename_date.isoformat()} 與資料夾日期 {expected_folder_date.isoformat()} 不一致",
            )
        )
    if expected_folder_date and order_date and order_date != expected_folder_date:
        issues.append(
            ImportIssue(
                path.name,
                None,
                "警告",
                f"Excel 日期 {order_date.isoformat()} 與資料夾日期 {expected_folder_date.isoformat()} 不一致",
            )
        )
    if meta.filename_customer and customer and meta.filename_customer != customer:
        issues.append(
            ImportIssue(
                path.name,
                None,
                "警告",
                f"檔名客戶「{meta.filename_customer}」與 Excel 客戶「{customer}」不一致",
            )
        )

    if order_status == "取消":
        issues.append(ImportIssue(path.name, None, "提示", "訂單狀態為取消，已保留記錄但不計入銷售數量及收入"))
        return [], issues
    if order_status not in COUNTED_ORDER_STATUSES:
        return [], issues

    # Product entry rows start after the top order header. Rows above 7 contain
    # customer/date/order fields, which must not be interpreted as item prices.
    for row_number in range(7, sheet.max_row + 1):
        values = [sheet.cell(row_number, col).value for col in range(1, 9)]
        text_values = [cell_text(value) for value in values]
        if not any(text_values):
            continue
        if is_section_or_header(text_values):
            continue

        quantity = to_decimal(values[0])
        product = text_values[2]
        category = text_values[3]
        part = text_values[4]
        # F 欄是所有分類適用的平均售價。
        average_price = to_decimal(values[5])
        buy_in_price = None
        cheap_price = to_decimal(values[6])
        fine_price = to_decimal(values[7])
        has_any_entry = any(value is not None for value in [quantity, values[5], cheap_price, fine_price])

        if not product and has_any_entry:
            issues.append(ImportIssue(path.name, row_number, "錯誤", "此行有数量/價錢，但沒有貨品名稱"))
            continue
        if not product:
            continue
        if master_products is not None and master_products:
            master_product = master_products.get(product)
            if master_product is None:
                issues.append(ImportIssue(path.name, row_number, "警告", f"{product}: 不在貨品清單內"))
            else:
                if not master_product.active:
                    issues.append(ImportIssue(path.name, row_number, "警告", f"{product}: 貨品清單狀態為停用"))
                if master_product.category and category and master_product.category != category:
                    issues.append(
                        ImportIssue(
                            path.name,
                            row_number,
                            "警告",
                            f"{product}: Excel 分類「{category}」與貨品清單「{master_product.category}」不一致",
                        )
                    )
        if quantity is None and not average_price and not cheap_price and not fine_price:
            continue

        if quantity is None:
            issues.append(ImportIssue(path.name, row_number, "錯誤", f"{product}: 缺少售出数量"))
            continue
        if quantity <= 0:
            issues.append(ImportIssue(path.name, row_number, "錯誤", f"{product}: 售出数量必須大於 0"))
            continue
        entered_prices = [price for price in (average_price, cheap_price, fine_price) if price is not None]
        if len(entered_prices) > 1:
            issues.append(ImportIssue(path.name, row_number, "錯誤", f"{product}: 只可填寫一種售價"))
            continue
        if not entered_prices:
            issues.append(ImportIssue(path.name, row_number, "錯誤", f"{product}: 缺少售價"))
            continue

        if average_price is not None:
            unit_price = average_price
            price_type = PRICE_TYPES["average"]
        elif cheap_price is not None:
            unit_price = cheap_price
            price_type = PRICE_TYPES["cheap"]
        else:
            unit_price = fine_price
            price_type = PRICE_TYPES["fine"]
        if unit_price is None or unit_price <= 0:
            issues.append(ImportIssue(path.name, row_number, "錯誤", f"{product}: 售價必須大於 0"))
            continue

        if order_status == "退貨":
            quantity = -quantity

        revenue = money(quantity * unit_price)
        estimated_margin = money(quantity * (unit_price - buy_in_price)) if buy_in_price is not None else None

        rows.append(
            SaleRow(
                source_file=path.name,
                row_number=row_number,
                order_date=order_date,
                customer=customer,
                order_number=order_number,
                order_status=order_status,
                category=category,
                product=product,
                part=part,
                quantity=quantity,
                unit_price=unit_price,
                price_type=price_type,
                buy_in_price=buy_in_price,
                revenue=revenue or Decimal("0"),
                estimated_margin=estimated_margin,
            )
        )

    if not rows:
        issues.append(ImportIssue(path.name, None, "警告", "沒有找到已售出貨品"))

    return rows, issues


def detect_duplicate_orders(sale_rows: list[SaleRow]) -> list[ImportIssue]:
    issues: list[ImportIssue] = []
    rows_by_file: dict[str, list[SaleRow]] = defaultdict(list)
    for row in sale_rows:
        rows_by_file[row.source_file].append(row)

    seen_identity: dict[tuple, str] = {}
    seen_content: dict[tuple, str] = {}

    for source_file, rows in sorted(rows_by_file.items()):
        if not rows:
            continue
        first = rows[0]
        identity = (first.order_date, first.customer, first.order_number)
        if first.order_number:
            previous_file = seen_identity.get(identity)
            if previous_file:
                issues.append(
                    ImportIssue(
                        source_file,
                        None,
                        "警告",
                        f"重複訂單編號：與 {previous_file} 使用相同日期、客戶及發貨單編號",
                    )
                )
            else:
                seen_identity[identity] = source_file

        content_lines = tuple(
            sorted(
                (
                    row.category,
                    row.product,
                    str(row.quantity),
                    str(row.unit_price),
                    row.price_type,
                )
                for row in rows
            )
        )
        content_signature = (first.order_date, first.customer, content_lines)
        previous_content_file = seen_content.get(content_signature)
        if previous_content_file and previous_content_file != seen_identity.get(identity):
            issues.append(
                ImportIssue(
                    source_file,
                    None,
                    "警告",
                    f"疑似重複訂單內容：與 {previous_content_file} 的客戶、日期、貨品、数量及售價相同",
                )
            )
        elif not previous_content_file:
            seen_content[content_signature] = source_file

    return issues


def detect_order_adjustments(sale_rows: list[SaleRow]) -> list[ImportIssue]:
    issues: list[ImportIssue] = []
    rows_by_file: dict[str, list[SaleRow]] = defaultdict(list)
    for row in sale_rows:
        rows_by_file[row.source_file].append(row)

    for source_file, rows in sorted(rows_by_file.items()):
        if not rows:
            continue
        status = rows[0].order_status
        if status == "更正":
            issues.append(ImportIssue(source_file, None, "提示", "訂單狀態為更正，請留意此單為修改後記錄"))
        elif status == "退貨":
            issues.append(ImportIssue(source_file, None, "提示", "訂單狀態為退貨，數量及收入已用負數計算"))
    return issues


def adjustment_counts(sale_rows: list[SaleRow], issues: list[ImportIssue]) -> dict[str, int]:
    return {
        "取消": len({issue.source_file for issue in issues if "訂單狀態為取消" in issue.message}),
        "更正": len({row.source_file for row in sale_rows if row.order_status == "更正"}),
        "退貨": len({row.source_file for row in sale_rows if row.order_status == "退貨"}),
    }


def grouped_values(rows: list[SaleRow], key_fn):
    result = defaultdict(lambda: {"quantity": Decimal("0"), "revenue": Decimal("0"), "files": set(), "rows": 0})
    for row in rows:
        key = key_fn(row)
        result[key]["quantity"] += row.quantity
        result[key]["revenue"] += row.revenue
        result[key]["files"].add(row.source_file)
        result[key]["rows"] += 1
    return result


def change_ratio(current: Decimal, previous: Decimal) -> Decimal | None:
    if previous == 0:
        return None
    return (current - previous) / previous * Decimal("100")


def format_percent(value: Decimal | None) -> str:
    if value is None:
        return "沒有上月資料"
    rounded = money(value) or Decimal("0")
    return f"{rounded:+.2f}%"


def append_alert(alerts: list[list[str]], severity: str, title: str, detail: str, action: str) -> None:
    alerts.append([severity, title, detail, action])


def build_daily_alerts(sale_rows: list[SaleRow], issues: list[ImportIssue]) -> list[list[str]]:
    alerts: list[list[str]] = []
    total_revenue = sum((row.revenue for row in sale_rows), Decimal("0"))
    adjustments = adjustment_counts(sale_rows, issues)

    if not sale_rows:
        append_alert(alerts, "高", "沒有有效銷售資料", "今日沒有讀到可計算的售出貨品。", "檢查訂單資料夾及 Excel 輸入。")
    if issues:
        append_alert(alerts, "中", "有輸入問題需要檢查", f"今日共有 {len(issues)} 個提示/警告/錯誤。", "查看「問題」分頁。")
    if any(adjustments.values()):
        append_alert(
            alerts,
            "中",
            "今日有訂單調整",
            f"取消 {adjustments['取消']} 張，更正 {adjustments['更正']} 張，退貨 {adjustments['退貨']} 張。",
            "確認調整原因及是否已通知相關人員。",
        )

    customers = grouped_values(sale_rows, lambda row: row.customer)
    if total_revenue > 0 and customers:
        top_customer, top_value = max(customers.items(), key=lambda item: item[1]["revenue"])
        share = top_value["revenue"] / total_revenue * Decimal("100")
        if share >= Decimal("60"):
            append_alert(
                alerts,
                "低",
                "收入集中在單一客戶",
                f"{top_customer} 佔今日收入 {format_percent(share)}。",
                "留意是否過度依賴單一客戶。",
            )

    products = grouped_values(sale_rows, lambda row: (row.category, row.product))
    if total_revenue > 0 and products:
        top_product, top_value = max(products.items(), key=lambda item: item[1]["revenue"])
        share = top_value["revenue"] / total_revenue * Decimal("100")
        if share >= Decimal("50"):
            append_alert(
                alerts,
                "低",
                "收入集中在單一貨品",
                f"{top_product[1]} 佔今日收入 {format_percent(share)}。",
                "留意此貨品供應及價格變化。",
            )

    if not alerts:
        append_alert(alerts, "正常", "沒有明顯異常", "今日銷售資料沒有觸發管理提示。", "可正常查看報表。")
    return alerts


def build_monthly_alerts(sale_rows: list[SaleRow], previous_rows: list[SaleRow], issues: list[ImportIssue]) -> list[list[str]]:
    alerts: list[list[str]] = []
    total_revenue = sum((row.revenue for row in sale_rows), Decimal("0"))
    previous_revenue = sum((row.revenue for row in previous_rows), Decimal("0"))
    total_quantity = sum((row.quantity for row in sale_rows), Decimal("0"))
    previous_quantity = sum((row.quantity for row in previous_rows), Decimal("0"))
    order_count = len({row.source_file for row in sale_rows})
    previous_order_count = len({row.source_file for row in previous_rows})
    adjustments = adjustment_counts(sale_rows, issues)

    if not sale_rows:
        append_alert(alerts, "高", "本月沒有有效銷售資料", "系統沒有讀到本月可計算的售出貨品。", "檢查本月每日訂單資料夾。")
    if issues:
        append_alert(alerts, "中", "本月有輸入問題需要檢查", f"共有 {len(issues)} 個提示/警告/錯誤。", "查看「問題」分頁並修正來源訂單。")
    if any(adjustments.values()):
        append_alert(
            alerts,
            "中",
            "本月有訂單調整",
            f"取消 {adjustments['取消']} 張，更正 {adjustments['更正']} 張，退貨 {adjustments['退貨']} 張。",
            "確認調整是否合理，並保留原始訂單記錄。",
        )

    revenue_change = change_ratio(total_revenue, previous_revenue)
    if revenue_change is not None:
        if revenue_change <= Decimal("-20"):
            append_alert(alerts, "高", "收入明顯下降", f"本月收入較上月 {format_percent(revenue_change)}。", "檢查主要客戶及主要貨品是否下跌。")
        elif revenue_change >= Decimal("20"):
            append_alert(alerts, "低", "收入明顯上升", f"本月收入較上月 {format_percent(revenue_change)}。", "確認增長來自哪些客戶/貨品。")

    quantity_change = change_ratio(total_quantity, previous_quantity)
    if quantity_change is not None:
        if quantity_change <= Decimal("-20"):
            append_alert(alerts, "高", "銷量明顯下降", f"本月售出数量較上月 {format_percent(quantity_change)}。", "檢查需求下跌或缺貨情況。")
        elif quantity_change >= Decimal("20"):
            append_alert(alerts, "低", "銷量明顯上升", f"本月售出数量較上月 {format_percent(quantity_change)}。", "留意庫存及供應是否足夠。")

    order_change = change_ratio(Decimal(order_count), Decimal(previous_order_count))
    if order_change is not None and order_change <= Decimal("-20"):
        append_alert(alerts, "中", "訂單數量下降", f"本月訂單數量較上月 {format_percent(order_change)}。", "檢查是否有客戶少下單。")

    current_customers = grouped_values(sale_rows, lambda row: row.customer)
    previous_customers = grouped_values(previous_rows, lambda row: row.customer)
    lost_customers = [
        customer
        for customer, previous in previous_customers.items()
        if previous["revenue"] > 0 and current_customers[customer]["revenue"] == 0
    ]
    if lost_customers:
        sample = "、".join(lost_customers[:5])
        append_alert(alerts, "高", "有客戶本月沒有再下單", f"{len(lost_customers)} 位上月客戶本月沒有收入：{sample}", "跟進是否流失、休業或資料未輸入。")

    dropped_customers = []
    for customer, previous in previous_customers.items():
        previous_value = previous["revenue"]
        current_value = current_customers[customer]["revenue"]
        ratio = change_ratio(current_value, previous_value)
        if ratio is not None and previous_value > 0 and current_value > 0 and ratio <= Decimal("-30"):
            dropped_customers.append((customer, ratio))
    if dropped_customers:
        customer, ratio = sorted(dropped_customers, key=lambda item: item[1])[0]
        append_alert(alerts, "中", "主要客戶收入下跌", f"{customer} 較上月 {format_percent(ratio)}。", "查看客戶分析，確認下跌原因。")

    current_products = grouped_values(sale_rows, lambda row: (row.category, row.product))
    previous_products = grouped_values(previous_rows, lambda row: (row.category, row.product))
    changed_products = []
    for product, previous in previous_products.items():
        previous_qty = previous["quantity"]
        current_qty = current_products[product]["quantity"]
        ratio = change_ratio(current_qty, previous_qty)
        if ratio is not None and previous_qty > 0 and abs(ratio) >= Decimal("30"):
            changed_products.append((product, ratio))
    if changed_products:
        product, ratio = sorted(changed_products, key=lambda item: abs(item[1]), reverse=True)[0]
        append_alert(alerts, "中", "貨品需求變化明顯", f"{product[1]} 售出数量較上月 {format_percent(ratio)}。", "查看貨品分析，調整採購或供應安排。")

    if total_revenue > 0 and current_customers:
        customer_revenues = sorted(current_customers.items(), key=lambda item: item[1]["revenue"], reverse=True)
        top_share = customer_revenues[0][1]["revenue"] / total_revenue * Decimal("100")
        if top_share >= Decimal("50"):
            append_alert(alerts, "中", "收入集中在單一客戶", f"{customer_revenues[0][0]} 佔本月收入 {format_percent(top_share)}。", "留意客戶集中風險。")

    if not alerts:
        append_alert(alerts, "正常", "沒有明顯異常", "本月銷售資料沒有觸發管理提示。", "可正常查看月報。")
    return alerts


def clamp_decimal(value: Decimal, minimum: Decimal, maximum: Decimal) -> Decimal:
    return max(minimum, min(maximum, value))


def forecast_confidence(elapsed_days: int, order_count: int, previous_rows: list[SaleRow]) -> str:
    if elapsed_days >= 14 and order_count >= 20 and previous_rows:
        return "高"
    if elapsed_days >= 7 and order_count >= 5:
        return "中"
    return "低"


def forecast_value(current_month_end: Decimal, previous_value: Decimal) -> Decimal:
    if previous_value <= 0:
        return current_month_end
    trend = (current_month_end - previous_value) / previous_value
    capped_trend = clamp_decimal(trend, Decimal("-0.30"), Decimal("0.30"))
    # Use half the trend to avoid overreacting to one abnormal month.
    return current_month_end * (Decimal("1") + capped_trend / Decimal("2"))


def build_forecast_rows(month: str, sale_rows: list[SaleRow], previous_rows: list[SaleRow]) -> list[list]:
    dates = [row.order_date for row in sale_rows if row.order_date]
    year = int(month[:4])
    month_number = int(month[4:])
    days_in_month = calendar.monthrange(year, month_number)[1]
    elapsed_days = max((day.day for day in dates), default=0)
    elapsed_days = max(1, min(elapsed_days, days_in_month))
    scale = Decimal(days_in_month) / Decimal(elapsed_days)
    order_count = len({row.source_file for row in sale_rows})
    previous_order_count = len({row.source_file for row in previous_rows})
    confidence = forecast_confidence(elapsed_days, order_count, previous_rows)

    total_revenue = sum((row.revenue for row in sale_rows), Decimal("0"))
    total_quantity = sum((row.quantity for row in sale_rows), Decimal("0"))
    previous_revenue = sum((row.revenue for row in previous_rows), Decimal("0"))
    previous_quantity = sum((row.quantity for row in previous_rows), Decimal("0"))

    forecast_rows = [
        [
            "總覽",
            "總收入 HKD",
            float(money(total_revenue) or Decimal("0")),
            float(money(total_revenue * scale) or Decimal("0")),
            float(money(forecast_value(total_revenue * scale, previous_revenue)) or Decimal("0")),
            confidence,
            f"根據本月首 {elapsed_days} 日資料估算；新系統初期只作參考。",
        ],
        [
            "總覽",
            "總售出数量",
            float(money(total_quantity) or Decimal("0")),
            float(money(total_quantity * scale) or Decimal("0")),
            float(money(forecast_value(total_quantity * scale, previous_quantity)) or Decimal("0")),
            confidence,
            f"根據本月首 {elapsed_days} 日資料估算；會受假期及大客戶訂單影響。",
        ],
        [
            "總覽",
            "總訂單數量",
            order_count,
            float(money(Decimal(order_count) * scale) or Decimal("0")),
            float(money(forecast_value(Decimal(order_count) * scale, Decimal(previous_order_count))) or Decimal("0")),
            confidence,
            "用目前每日訂單速度估算月底及下月訂單量。",
        ],
    ]

    current_products = grouped_values(sale_rows, lambda row: (row.category, row.product))
    previous_products = grouped_values(previous_rows, lambda row: (row.category, row.product))
    for (category, product), current in sorted(current_products.items(), key=lambda item: item[1]["revenue"], reverse=True)[:8]:
        previous = previous_products[(category, product)]
        current_quantity = current["quantity"]
        current_revenue = current["revenue"]
        month_end_quantity = current_quantity * scale
        month_end_revenue = current_revenue * scale
        next_quantity = forecast_value(month_end_quantity, previous["quantity"])
        next_revenue = forecast_value(month_end_revenue, previous["revenue"])
        forecast_rows.append(
            [
                "貨品",
                f"{category} - {product}",
                f"数量 {float(money(current_quantity) or Decimal('0'))} / HKD {float(money(current_revenue) or Decimal('0')):,.2f}",
                f"数量 {float(money(month_end_quantity) or Decimal('0'))} / HKD {float(money(month_end_revenue) or Decimal('0')):,.2f}",
                f"数量 {float(money(next_quantity) or Decimal('0'))} / HKD {float(money(next_revenue) or Decimal('0')):,.2f}",
                confidence,
                "貨品預測只列本月收入最高貨品；用於採購及供應參考。",
            ]
        )

    if not sale_rows:
        forecast_rows.append(["提示", "沒有足夠資料", "", "", "", "低", "本月未有有效銷售資料，暫時不能預測。"])
    return forecast_rows


def safe_table_name(title: str) -> str:
    # Excel table names are strict; use ASCII regardless of sheet title.
    suffix = abs(hash(title)) % 1_000_000
    return f"Table_{suffix}"


def add_sheet(workbook: Workbook, title: str, headers: list[str], data: list[list], make_table: bool = True):
    sheet = workbook.create_sheet(title)
    sheet.append(headers)
    for row in data:
        sheet.append(row)
    style_header(sheet)
    if make_table and headers:
        end_row = max(2, len(data) + 1)
        end_col = len(headers)
        table = Table(displayName=safe_table_name(title), ref=f"A1:{get_column_letter(end_col)}{end_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)
    autosize(sheet)
    return sheet


def style_header(sheet):
    fill = PatternFill("solid", fgColor="115E59")
    font = Font(bold=True, color="FFFFFF")
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 34
    sheet.freeze_panes = "A2"


def autosize(sheet):
    for column in sheet.columns:
        width = 10
        letter = get_column_letter(column[0].column)
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            display_width = sum(2 if unicodedata.east_asian_width(char) in {"W", "F", "A"} else 1 for char in value)
            width = max(width, min(display_width + 3, 48))
        sheet.column_dimensions[letter].width = width


def append_sheet_section(target, source, section_title: str) -> None:
    start_row = target.max_row + 3
    target.cell(start_row, 1, section_title)
    target.cell(start_row, 1).font = Font(bold=True, size=14, color="115E59")
    for source_row in source.iter_rows(values_only=True):
        target.append(list(source_row))
    header_row = start_row + 1
    fill = PatternFill("solid", fgColor="115E59")
    font = Font(bold=True, color="FFFFFF")
    for cell in target[header_row]:
        if cell.value is not None:
            cell.fill = fill
            cell.font = font
    autosize(target)


def consolidate_monthly_sheets(workbook: Workbook) -> None:
    customer_sheet = workbook["客戶月報"]
    product_sheet = workbook["貨品月報"]
    category_sheet = workbook["分類月報"]
    issue_sheet = workbook["問題"]

    append_sheet_section(customer_sheet, workbook["客戶比較"], "客戶本月與上月比較")
    append_sheet_section(product_sheet, workbook["貨品比較"], "貨品本月與上月比較")
    append_sheet_section(category_sheet, workbook["分類比較"], "分類本月與上月比較")
    append_sheet_section(issue_sheet, workbook["問題摘要"], "問題摘要")

    customer_sheet.title = "客戶分析"
    product_sheet.title = "貨品分析"
    category_sheet.title = "分類分析"

    for sheet_name in ["客戶比較", "貨品比較", "分類比較", "十大客戶", "十大貨品", "問題摘要"]:
        workbook[sheet_name].sheet_state = "hidden"


def add_monthly_charts(
    workbook: Workbook,
    summary_sheet,
    category_sheet,
    trend_sheet,
    top_customer_sheet,
    top_product_sheet,
    comparison_sheet=None,
):
    dashboard = summary_sheet

    if category_sheet.max_row >= 2:
        pie = PieChart()
        labels = Reference(category_sheet, min_col=1, min_row=2, max_row=category_sheet.max_row)
        data = Reference(category_sheet, min_col=4, min_row=1, max_row=category_sheet.max_row)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = "分類收入佔比"
        pie.height = 8
        pie.width = 10
        dashboard.add_chart(pie, "D2")

    if trend_sheet.max_row >= 2:
        line = LineChart()
        data = Reference(trend_sheet, min_col=5, min_row=1, max_row=trend_sheet.max_row)
        labels = Reference(trend_sheet, min_col=1, min_row=2, max_row=trend_sheet.max_row)
        line.add_data(data, titles_from_data=True)
        line.set_categories(labels)
        line.title = "每日收入走勢"
        line.y_axis.title = "收入 HKD"
        line.x_axis.title = "日期"
        line.height = 8
        line.width = 16
        dashboard.add_chart(line, "N2")

    if top_customer_sheet.max_row >= 2:
        bar = BarChart()
        data = Reference(top_customer_sheet, min_col=2, min_row=1, max_row=top_customer_sheet.max_row)
        labels = Reference(top_customer_sheet, min_col=1, min_row=2, max_row=top_customer_sheet.max_row)
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(labels)
        bar.title = "十大客戶收入"
        bar.y_axis.title = "收入 HKD"
        bar.height = 8
        bar.width = 16
        dashboard.add_chart(bar, "D20")

    if top_product_sheet.max_row >= 2:
        bar = BarChart()
        data = Reference(top_product_sheet, min_col=4, min_row=1, max_row=top_product_sheet.max_row)
        labels = Reference(top_product_sheet, min_col=2, min_row=2, max_row=top_product_sheet.max_row)
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(labels)
        bar.title = "十大貨品收入"
        bar.y_axis.title = "收入 HKD"
        bar.height = 8
        bar.width = 16
        dashboard.add_chart(bar, "N20")

    if comparison_sheet is not None and comparison_sheet.max_row >= 2:
        comparison = BarChart()
        data = Reference(comparison_sheet, min_col=2, max_col=3, min_row=1, max_row=comparison_sheet.max_row)
        labels = Reference(comparison_sheet, min_col=1, min_row=2, max_row=comparison_sheet.max_row)
        comparison.add_data(data, titles_from_data=True)
        comparison.set_categories(labels)
        comparison.title = "本月與上月比較"
        comparison.height = 9
        comparison.width = 18
        dashboard.add_chart(comparison, "D38")

    autosize(dashboard)


def write_report(output_path: Path, sale_rows: list[SaleRow], issues: list[ImportIssue]):
    workbook = Workbook()
    workbook.remove(workbook.active)

    total_revenue = sum((row.revenue for row in sale_rows), Decimal("0"))
    total_quantity = sum((row.quantity for row in sale_rows), Decimal("0"))
    dates = sorted({row.order_date for row in sale_rows if row.order_date})
    report_date = dates[0].isoformat() if len(dates) == 1 else "Multiple / unknown"
    adjustments = adjustment_counts(sale_rows, issues)

    summary = [
        ["報表日期", report_date],
        ["總訂單數量", len({row.source_file for row in sale_rows})],
        ["售出貨品記錄數", len(sale_rows)],
        ["總售出数量", float(total_quantity)],
        ["總收入 HKD", float(total_revenue)],
        ["取消訂單數量", adjustments["取消"]],
        ["更正訂單數量", adjustments["更正"]],
        ["退貨訂單數量", adjustments["退貨"]],
        ["問題數量", len(issues)],
    ]
    add_sheet(workbook, "總覽", ["項目", "數值"], summary)
    add_sheet(workbook, "管理提示", ["嚴重程度", "提示", "內容", "建議行動"], build_daily_alerts(sale_rows, issues))

    grouped = defaultdict(lambda: {"quantity": Decimal("0"), "revenue": Decimal("0"), "rows": 0})
    for row in sale_rows:
        key = row.customer
        grouped[key]["quantity"] += row.quantity
        grouped[key]["revenue"] += row.revenue
        grouped[key]["rows"] += 1
    add_sheet(
        workbook,
        "按客戶",
        ["客戶", "售出貨品記錄數", "售出数量", "收入 HKD"],
        [[key, val["rows"], float(val["quantity"]), float(val["revenue"])] for key, val in sorted(grouped.items())],
    )

    grouped.clear()
    for row in sale_rows:
        key = (row.category, row.product)
        grouped[key]["quantity"] += row.quantity
        grouped[key]["revenue"] += row.revenue
        grouped[key]["rows"] += 1
    add_sheet(
        workbook,
        "按貨品",
        ["分類", "貨品", "售出貨品記錄數", "售出数量", "收入 HKD"],
        [[cat, product, val["rows"], float(val["quantity"]), float(val["revenue"])] for (cat, product), val in sorted(grouped.items())],
    )

    grouped.clear()
    for row in sale_rows:
        grouped[row.category]["quantity"] += row.quantity
        grouped[row.category]["revenue"] += row.revenue
        grouped[row.category]["rows"] += 1
    add_sheet(
        workbook,
        "按分類",
        ["分類", "售出貨品記錄數", "售出数量", "收入 HKD"],
        [[key, val["rows"], float(val["quantity"]), float(val["revenue"])] for key, val in sorted(grouped.items())],
    )

    add_sheet(
        workbook,
        "明細",
        [
            "來源檔案",
            "日期",
            "客戶",
            "發貨單編號",
            "訂單狀態",
            "分類",
            "貨品",
            "部位",
            "售出数量",
            "售價",
            "售價類型",
            "平均買入價",
            "收入",
        ],
        [
            [
                row.source_file,
                row.order_date.isoformat() if row.order_date else "",
                row.customer,
                row.order_number,
                row.order_status,
                row.category,
                row.product,
                row.part,
                float(row.quantity),
                float(row.unit_price),
                row.price_type,
                float(row.buy_in_price) if row.buy_in_price is not None else "",
                float(row.revenue),
            ]
            for row in sale_rows
        ],
    )

    add_sheet(
        workbook,
        "問題",
        ["來源檔案", "嚴重程度", "問題"],
        [[issue.source_file, issue.severity, issue.message] for issue in issues],
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def percent_change(current: Decimal, previous: Decimal):
    if previous == 0:
        return "沒有上月資料" if current != 0 else 0
    return float(money((current - previous) / previous * Decimal("100")) or Decimal("0"))


def write_monthly_report(
    output_path: Path,
    month: str,
    sale_rows: list[SaleRow],
    issues: list[ImportIssue],
    previous_rows: list[SaleRow] | None = None,
):
    previous_rows = previous_rows or []
    workbook = Workbook()
    workbook.remove(workbook.active)

    total_revenue = sum((row.revenue for row in sale_rows), Decimal("0"))
    total_quantity = sum((row.quantity for row in sale_rows), Decimal("0"))
    order_files = len({row.source_file for row in sale_rows})
    customers = len({row.customer for row in sale_rows if row.customer})
    avg_order = money(total_revenue / Decimal(order_files)) if order_files else Decimal("0")
    avg_price = money(total_revenue / total_quantity) if total_quantity else Decimal("0")
    adjustments = adjustment_counts(sale_rows, issues)

    summary = [
        ["報表月份", month],
        ["總訂單數量", order_files],
        ["客戶數量", customers],
        ["售出貨品記錄數", len(sale_rows)],
        ["總售出数量", float(total_quantity)],
        ["總收入 HKD", float(total_revenue)],
        ["平均每單收入 HKD", float(avg_order or Decimal("0"))],
        ["平均售價 HKD", float(avg_price or Decimal("0"))],
        ["取消訂單數量", adjustments["取消"]],
        ["更正訂單數量", adjustments["更正"]],
        ["退貨訂單數量", adjustments["退貨"]],
        ["問題數量", len(issues)],
    ]
    summary_sheet = add_sheet(workbook, "總覽", ["項目", "數值"], summary)
    add_sheet(workbook, "管理提示", ["嚴重程度", "提示", "內容", "建議行動"], build_monthly_alerts(sale_rows, previous_rows, issues))
    add_sheet(
        workbook,
        "銷售預測",
        ["類型", "項目", "目前資料", "本月月底預測", "下月預測", "信心", "說明"],
        build_forecast_rows(month, sale_rows, previous_rows),
    )

    grouped = defaultdict(lambda: {"quantity": Decimal("0"), "revenue": Decimal("0"), "rows": 0, "files": set()})
    for row in sale_rows:
        grouped[row.customer]["quantity"] += row.quantity
        grouped[row.customer]["revenue"] += row.revenue
        grouped[row.customer]["rows"] += 1
        grouped[row.customer]["files"].add(row.source_file)
    customer_sheet = add_sheet(
        workbook,
        "客戶月報",
        ["客戶", "總訂單數量", "售出貨品記錄數", "售出数量", "收入 HKD", "平均售價 HKD", "收入佔比 %"],
        [
            [
                key,
                len(val["files"]),
                val["rows"],
                float(val["quantity"]),
                float(val["revenue"]),
                float(money(val["revenue"] / val["quantity"]) or Decimal("0")) if val["quantity"] else "",
                float(money(val["revenue"] / total_revenue * Decimal("100")) or Decimal("0")) if total_revenue else 0,
            ]
            for key, val in sorted(grouped.items(), key=lambda item: item[1]["revenue"], reverse=True)
        ],
    )

    grouped.clear()
    for row in sale_rows:
        key = (row.category, row.product)
        grouped[key]["quantity"] += row.quantity
        grouped[key]["revenue"] += row.revenue
        grouped[key]["rows"] += 1
    product_sheet = add_sheet(
        workbook,
        "貨品月報",
        ["分類", "貨品", "售出貨品記錄數", "售出数量", "收入 HKD", "平均售價 HKD", "收入佔比 %"],
        [
            [
                cat,
                product,
                val["rows"],
                float(val["quantity"]),
                float(val["revenue"]),
                float(money(val["revenue"] / val["quantity"]) or Decimal("0")) if val["quantity"] else "",
                float(money(val["revenue"] / total_revenue * Decimal("100")) or Decimal("0")) if total_revenue else 0,
            ]
            for (cat, product), val in sorted(grouped.items(), key=lambda item: item[1]["revenue"], reverse=True)
        ],
    )

    grouped.clear()
    for row in sale_rows:
        grouped[row.category]["quantity"] += row.quantity
        grouped[row.category]["revenue"] += row.revenue
        grouped[row.category]["rows"] += 1
    category_sheet = add_sheet(
        workbook,
        "分類月報",
        ["分類", "售出貨品記錄數", "售出数量", "收入 HKD", "收入佔比 %"],
        [
            [
                key,
                val["rows"],
                float(val["quantity"]),
                float(val["revenue"]),
                float(money(val["revenue"] / total_revenue * Decimal("100")) or Decimal("0")) if total_revenue else 0,
            ]
            for key, val in sorted(grouped.items(), key=lambda item: item[1]["revenue"], reverse=True)
        ],
    )

    grouped.clear()
    for row in sale_rows:
        key = row.order_date.isoformat() if row.order_date else ""
        grouped[key]["quantity"] += row.quantity
        grouped[key]["revenue"] += row.revenue
        grouped[key]["rows"] += 1
        grouped[key]["files"].add(row.source_file)
    trend_sheet = add_sheet(
        workbook,
        "每日走勢",
        ["日期", "總訂單數量", "售出貨品記錄數", "售出数量", "收入 HKD"],
        [
            [key, len(val["files"]), val["rows"], float(val["quantity"]), float(val["revenue"])]
            for key, val in sorted(grouped.items())
        ],
    )

    customer_revenue = defaultdict(lambda: Decimal("0"))
    for row in sale_rows:
        customer_revenue[row.customer] += row.revenue
    top_customer_sheet = add_sheet(
        workbook,
        "十大客戶",
        ["客戶", "收入 HKD", "收入佔比 %"],
        [
            [
                customer,
                float(revenue),
                float(money(revenue / total_revenue * Decimal("100")) or Decimal("0")) if total_revenue else 0,
            ]
            for customer, revenue in sorted(customer_revenue.items(), key=lambda item: item[1], reverse=True)[:10]
        ],
    )

    product_revenue = defaultdict(lambda: {"quantity": Decimal("0"), "revenue": Decimal("0")})
    for row in sale_rows:
        key = (row.category, row.product)
        product_revenue[key]["quantity"] += row.quantity
        product_revenue[key]["revenue"] += row.revenue
    top_product_sheet = add_sheet(
        workbook,
        "十大貨品",
        ["分類", "貨品", "售出数量", "收入 HKD", "平均售價 HKD"],
        [
            [
                category,
                product,
                float(val["quantity"]),
                float(val["revenue"]),
                float(money(val["revenue"] / val["quantity"]) or Decimal("0")) if val["quantity"] else "",
            ]
            for (category, product), val in sorted(product_revenue.items(), key=lambda item: item[1]["revenue"], reverse=True)[:10]
        ],
    )

    issue_group = defaultdict(int)
    for issue in issues:
        issue_group[issue.severity] += 1
    issue_summary_sheet = add_sheet(
        workbook,
        "問題摘要",
        ["嚴重程度", "數量"],
        [[severity, count] for severity, count in sorted(issue_group.items())],
    )

    previous_revenue = sum((row.revenue for row in previous_rows), Decimal("0"))
    previous_quantity = sum((row.quantity for row in previous_rows), Decimal("0"))
    previous_orders = len({row.source_file for row in previous_rows})
    previous_customers = len({row.customer for row in previous_rows if row.customer})
    current_avg_order = total_revenue / Decimal(order_files) if order_files else Decimal("0")
    previous_avg_order = previous_revenue / Decimal(previous_orders) if previous_orders else Decimal("0")

    comparison_sheet = add_sheet(
        workbook,
        "月度比較",
        ["指標", "本月", "上月", "變化 %"],
        [
            ["總收入 HKD", float(total_revenue), float(previous_revenue), percent_change(total_revenue, previous_revenue)],
            ["總售出数量", float(total_quantity), float(previous_quantity), percent_change(total_quantity, previous_quantity)],
            ["總訂單數量", order_files, previous_orders, percent_change(Decimal(order_files), Decimal(previous_orders))],
            ["客戶數量", customers, previous_customers, percent_change(Decimal(customers), Decimal(previous_customers))],
            ["平均每單收入 HKD", float(money(current_avg_order) or Decimal("0")), float(money(previous_avg_order) or Decimal("0")), percent_change(current_avg_order, previous_avg_order)],
        ],
    )

    current_customer = grouped_values(sale_rows, lambda row: row.customer)
    previous_customer = grouped_values(previous_rows, lambda row: row.customer)
    customer_keys = sorted(set(current_customer) | set(previous_customer))
    add_sheet(
        workbook,
        "客戶比較",
        ["客戶", "本月收入", "上月收入", "收入變化 %", "本月数量", "上月数量", "数量變化 %"],
        [[key, float(current_customer[key]["revenue"]), float(previous_customer[key]["revenue"]), percent_change(current_customer[key]["revenue"], previous_customer[key]["revenue"]), float(current_customer[key]["quantity"]), float(previous_customer[key]["quantity"]), percent_change(current_customer[key]["quantity"], previous_customer[key]["quantity"])] for key in customer_keys],
    )

    current_product = grouped_values(sale_rows, lambda row: (row.category, row.product))
    previous_product = grouped_values(previous_rows, lambda row: (row.category, row.product))
    product_keys = sorted(set(current_product) | set(previous_product))
    add_sheet(
        workbook,
        "貨品比較",
        ["分類", "貨品", "本月收入", "上月收入", "收入變化 %", "本月数量", "上月数量", "数量變化 %"],
        [[key[0], key[1], float(current_product[key]["revenue"]), float(previous_product[key]["revenue"]), percent_change(current_product[key]["revenue"], previous_product[key]["revenue"]), float(current_product[key]["quantity"]), float(previous_product[key]["quantity"]), percent_change(current_product[key]["quantity"], previous_product[key]["quantity"])] for key in product_keys],
    )

    current_category = grouped_values(sale_rows, lambda row: row.category)
    previous_category = grouped_values(previous_rows, lambda row: row.category)
    category_keys = sorted(set(current_category) | set(previous_category))
    add_sheet(
        workbook,
        "分類比較",
        ["分類", "本月收入", "上月收入", "收入變化 %", "本月数量", "上月数量", "数量變化 %"],
        [[key, float(current_category[key]["revenue"]), float(previous_category[key]["revenue"]), percent_change(current_category[key]["revenue"], previous_category[key]["revenue"]), float(current_category[key]["quantity"]), float(previous_category[key]["quantity"]), percent_change(current_category[key]["quantity"], previous_category[key]["quantity"])] for key in category_keys],
    )

    add_monthly_charts(
        workbook,
        summary_sheet,
        category_sheet,
        trend_sheet,
        top_customer_sheet,
        top_product_sheet,
        comparison_sheet,
    )

    add_sheet(
        workbook,
        "明細",
        [
            "來源檔案",
            "日期",
            "客戶",
            "發貨單編號",
            "訂單狀態",
            "分類",
            "貨品",
            "部位",
            "售出数量",
            "售價",
            "售價類型",
            "平均買入價",
            "收入",
        ],
        [
            [
                row.source_file,
                row.order_date.isoformat() if row.order_date else "",
                row.customer,
                row.order_number,
                row.order_status,
                row.category,
                row.product,
                row.part,
                float(row.quantity),
                float(row.unit_price),
                row.price_type,
                float(row.buy_in_price) if row.buy_in_price is not None else "",
                float(row.revenue),
            ]
            for row in sale_rows
        ],
    )

    add_sheet(
        workbook,
        "問題",
        ["來源檔案", "嚴重程度", "問題"],
        [[issue.source_file, issue.severity, issue.message] for issue in issues],
    )

    consolidate_monthly_sheets(workbook)
    reorder_monthly_sheets(workbook)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def reorder_monthly_sheets(workbook: Workbook) -> None:
    preferred_order = [
        "總覽",
        "管理提示",
        "銷售預測",
        "月度比較",
        "客戶分析",
        "貨品分析",
        "分類分析",
        "每日走勢",
        "明細",
        "問題",
        "客戶比較",
        "貨品比較",
        "分類比較",
        "十大客戶",
        "十大貨品",
        "問題摘要",
    ]
    workbook._sheets = [workbook[name] for name in preferred_order if name in workbook.sheetnames]


def load_master_context() -> tuple[dict[str, MasterProduct], set[str]]:
    ensure_master_data()
    return load_master_products(MASTER_DATA_PATH), load_master_customers(MASTER_DATA_PATH)


def timestamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def copy_files_to_backup(files: list[Path], backup_root: Path) -> None:
    backup_root.mkdir(parents=True, exist_ok=True)
    for source in files:
        if not source.exists() or not source.is_file():
            continue
        target = backup_root / source.name
        shutil.copy2(source, target)


def backup_generated_report(output_path: Path, backup_root: Path) -> None:
    files = [output_path, output_path.with_suffix(".pdf")]
    copy_files_to_backup([path for path in files if path.exists()], backup_root)


def generate_daily_report(folder: Path, output_path: Path) -> tuple[int, int, int]:
    """Read a folder of order workbooks, write report, return counts.

    Returns:
        (file_count, sold_row_count, issue_count)
    """
    all_rows: list[SaleRow] = []
    all_issues: list[ImportIssue] = []
    workbook_paths = list(iter_workbooks(folder))
    expected_folder_date = folder_date(folder)
    master_products, master_customers = load_master_context()
    backup_root = output_path.parents[1] / "Backups" if len(output_path.parents) > 1 else output_path.parent / "Backups"
    run_backup = backup_root / f"daily_{folder.name}_{timestamp()}"
    copy_files_to_backup(workbook_paths, run_backup / "input_orders")

    for workbook_path in workbook_paths:
        rows, issues = read_order_file(workbook_path, expected_folder_date, master_products, master_customers)
        all_rows.extend(rows)
        all_issues.extend(issues)

    all_issues.extend(detect_duplicate_orders(all_rows))
    all_issues.extend(detect_order_adjustments(all_rows))
    write_report(output_path, all_rows, all_issues)
    from report_pdfs import write_daily_pdf

    write_daily_pdf(output_path.with_suffix(".pdf"), all_rows, all_issues)
    backup_generated_report(output_path, run_backup / "reports")
    return len(workbook_paths), len(all_rows), len(all_issues)


def generate_monthly_report(orders_base_folder: Path, month: str, output_path: Path) -> tuple[int, int, int]:
    """Read all YYYYMMDD folders for a month and generate a monthly report.

    Args:
        orders_base_folder: folder containing day folders like 20260618.
        month: YYYYMM.
        output_path: report path.
    """
    if not re.fullmatch(r"\d{6}", month):
        raise ValueError("month must be YYYYMM")

    all_rows: list[SaleRow] = []
    all_issues: list[ImportIssue] = []
    previous_rows: list[SaleRow] = []
    file_count = 0
    current_month_date = datetime.strptime(month + "01", "%Y%m%d").date()
    previous_month = (current_month_date.replace(day=1) - timedelta(days=1)).strftime("%Y%m")
    master_products, master_customers = load_master_context()
    backup_root = output_path.parents[1] / "Backups" if len(output_path.parents) > 1 else output_path.parent / "Backups"
    run_backup = backup_root / f"monthly_{month}_{timestamp()}"

    for day_folder in sorted(orders_base_folder.iterdir()):
        if not day_folder.is_dir():
            continue
        if not re.fullmatch(r"\d{8}", day_folder.name):
            continue
        folder_month = day_folder.name[:6]
        if folder_month not in {month, previous_month}:
            continue

        expected_folder_date = folder_date(day_folder)
        workbook_paths = list(iter_workbooks(day_folder))
        if folder_month == month:
            file_count += len(workbook_paths)
            copy_files_to_backup(workbook_paths, run_backup / "input_orders" / day_folder.name)
        for workbook_path in workbook_paths:
            rows, issues = read_order_file(workbook_path, expected_folder_date, master_products, master_customers)
            if folder_month == month:
                all_rows.extend(rows)
                all_issues.extend(issues)
            else:
                previous_rows.extend(rows)

    all_issues.extend(detect_duplicate_orders(all_rows))
    all_issues.extend(detect_order_adjustments(all_rows))
    write_monthly_report(output_path, month, all_rows, all_issues, previous_rows)
    from report_pdfs import write_monthly_pdf

    write_monthly_pdf(output_path.with_suffix(".pdf"), month, all_rows, previous_rows, all_issues)
    backup_generated_report(output_path, run_backup / "reports")
    return file_count, len(all_rows), len(all_issues)


def main() -> int:
    parser = argparse.ArgumentParser(description="Read order Excel files and generate a daily report.")
    parser.add_argument("folder", type=Path, help="Folder containing YYYYMMDD_客戶名_序號.xlsx files")
    parser.add_argument("--output", type=Path, default=None, help="Output .xlsx path")
    args = parser.parse_args()

    if not args.folder.exists() or not args.folder.is_dir():
        raise SystemExit(f"Folder does not exist: {args.folder}")

    if args.output:
        output_path = args.output
    else:
        output_path = args.folder / "daily_report.xlsx"

    file_count, sold_row_count, issue_count = generate_daily_report(args.folder, output_path)

    print(f"Read files: {file_count}")
    print(f"Sold rows: {sold_row_count}")
    print(f"Issues: {issue_count}")
    print(f"Report: {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
