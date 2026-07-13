from __future__ import annotations

import sys
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from copy import copy


PROJECT_ROOT = Path(sys.executable).resolve().parent if getattr(sys, "frozen", False) else Path(__file__).resolve().parents[1]
MASTER_DATA_PATH = PROJECT_ROOT / "Master Data" / "master-data.xlsx"
ORDER_TEMPLATE_PATH = PROJECT_ROOT / "Order Template" / "order-file-template.xlsx"
LEGACY_TEMPLATE_PATH = PROJECT_ROOT / "Daily Reports" / "order-file-template-updated.xlsx"
TEMPLATE_PATH = ORDER_TEMPLATE_PATH if ORDER_TEMPLATE_PATH.exists() else LEGACY_TEMPLATE_PATH
PACKAGE_TEMPLATE_PATH = PROJECT_ROOT / "Company Package" / "銷售報表系統" / "Order Template" / "order-file-template.xlsx"

CATEGORIES = ["牛", "豬", "魚", "羊", "雜貨"]
ORDER_STATUSES = ["正常", "取消", "更正", "退貨"]
PRODUCT_HEADERS = ["貨品", "分類", "部位", "單位", "狀態"]
CUSTOMER_HEADERS = ["客戶名稱", "狀態"]
PRODUCT_ROWS = [(9, 48), (53, 79), (84, 84), (89, 94)]
INPUT_ROWS = list(range(9, 49)) + list(range(53, 80)) + [84] + list(range(89, 95))
PRODUCT_TABLE_HEADERS = [
    "售出数量",
    "代碼",
    "品名",
    "分類",
    "部位",
    "平均售價\nHKD/KG",
    "平價切售價\nHKD/KG",
    "精修切售價\nHKD/KG",
]


@dataclass(frozen=True)
class MasterProduct:
    name: str
    category: str
    part: str
    unit: str
    active: bool


def cell_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def style_header(sheet) -> None:
    fill = PatternFill("solid", fgColor="115E59")
    font = Font(bold=True, color="FFFFFF")
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.freeze_panes = "A2"


def autosize(sheet) -> None:
    for column in sheet.columns:
        letter = get_column_letter(column[0].column)
        width = 10
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, min(len(value) * 2 + 2, 42))
        sheet.column_dimensions[letter].width = width


def add_table(sheet, name: str, max_col: int, max_row: int) -> None:
    if max_row < 2:
        max_row = 2
    table = Table(displayName=name, ref=f"A1:{get_column_letter(max_col)}{max_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)


def extract_products_from_template(template_path: Path = TEMPLATE_PATH) -> list[MasterProduct]:
    if not template_path.exists():
        return []
    workbook = load_workbook(template_path, data_only=False)
    sheet = workbook[workbook.sheetnames[0]]
    products: list[MasterProduct] = []
    seen: set[str] = set()
    for start, end in PRODUCT_ROWS:
        for row in range(start, end + 1):
            name = cell_text(sheet.cell(row, 3).value)
            category = cell_text(sheet.cell(row, 4).value)
            part = cell_text(sheet.cell(row, 5).value) or name
            if not name or name in seen:
                continue
            seen.add(name)
            products.append(MasterProduct(name=name, category=category, part=part, unit="KG", active=True))
    return products


def create_master_data(master_path: Path = MASTER_DATA_PATH, template_path: Path = TEMPLATE_PATH) -> Path:
    master_path.parent.mkdir(parents=True, exist_ok=True)
    products = extract_products_from_template(template_path)

    workbook = Workbook()
    product_sheet = workbook.active
    product_sheet.title = "貨品清單"
    product_sheet.append(PRODUCT_HEADERS)
    for product in products:
        product_sheet.append([product.name, product.category, product.part, product.unit, "使用中"])
    style_header(product_sheet)
    autosize(product_sheet)
    add_table(product_sheet, "ProductMaster", len(PRODUCT_HEADERS), max(2, product_sheet.max_row))

    customer_sheet = workbook.create_sheet("客戶清單")
    customer_sheet.append(CUSTOMER_HEADERS)
    customer_sheet.append(["請在這裡加入客戶名稱", "使用中"])
    style_header(customer_sheet)
    autosize(customer_sheet)
    add_table(customer_sheet, "CustomerMaster", len(CUSTOMER_HEADERS), max(2, customer_sheet.max_row))

    notes = workbook.create_sheet("說明")
    notes.append(["用途", "內容"])
    notes.append(["貨品清單", "新增可售貨品時，在貨品清單加入一行。分類請用：牛、豬、魚、羊、雜貨。"])
    notes.append(["客戶清單", "新增客戶時，在客戶清單加入一行。"])
    notes.append(["狀態", "使用中代表可選用；停用代表保留歷史資料，但不建議新訂單使用。"])
    style_header(notes)
    autosize(notes)

    workbook.save(master_path)
    return master_path


def ensure_master_data(master_path: Path = MASTER_DATA_PATH, template_path: Path = TEMPLATE_PATH) -> Path:
    if not master_path.exists():
        create_master_data(master_path, template_path)
    return master_path


def load_master_products(master_path: Path = MASTER_DATA_PATH) -> dict[str, MasterProduct]:
    if not master_path.exists():
        return {}
    workbook = load_workbook(master_path, data_only=True)
    if "貨品清單" not in workbook.sheetnames:
        return {}
    sheet = workbook["貨品清單"]
    products: dict[str, MasterProduct] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        name = cell_text(row[0] if len(row) > 0 else "")
        category = cell_text(row[1] if len(row) > 1 else "")
        part = cell_text(row[2] if len(row) > 2 else "") or name
        unit = cell_text(row[3] if len(row) > 3 else "") or "KG"
        status = cell_text(row[4] if len(row) > 4 else "") or "使用中"
        if not name:
            continue
        products[name] = MasterProduct(
            name=name,
            category=category,
            part=part,
            unit=unit,
            active=status != "停用",
        )
    return products


def load_master_customers(master_path: Path = MASTER_DATA_PATH) -> set[str]:
    if not master_path.exists():
        return set()
    workbook = load_workbook(master_path, data_only=True)
    if "客戶清單" not in workbook.sheetnames:
        return set()
    sheet = workbook["客戶清單"]
    customers: set[str] = set()
    for row in sheet.iter_rows(min_row=2, values_only=True):
        name = cell_text(row[0] if row else "")
        status = cell_text(row[1] if len(row) > 1 else "") or "使用中"
        if name and status != "停用" and not name.startswith("請在這裡"):
            customers.add(name)
    return customers


def copy_row_style(sheet, source_row: int, target_row: int, max_col: int = 8) -> None:
    source_height = sheet.row_dimensions[source_row].height
    if source_height:
        sheet.row_dimensions[target_row].height = source_height
    for col in range(1, max_col + 1):
        source = sheet.cell(source_row, col)
        target = sheet.cell(target_row, col)
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.number_format = source.number_format
        target.protection = copy(source.protection)


def clear_product_area(sheet) -> None:
    for merged_range in list(sheet.merged_cells.ranges):
        if merged_range.min_row >= 7:
            sheet.unmerge_cells(str(merged_range))
    if sheet.max_row >= 7:
        sheet.delete_rows(7, sheet.max_row - 6)


def rebuild_product_area(sheet, products: list[MasterProduct]) -> list[int]:
    section_style_row = 7
    header_style_row = 8
    product_style_row = 9
    summary_title_style_row = 97
    summary_header_style_row = 98
    summary_value_style_row = 99

    section_styles = []
    for source_row in [section_style_row, header_style_row, product_style_row, summary_title_style_row, summary_header_style_row, summary_value_style_row]:
        row_styles = []
        for col in range(1, 9):
            source = sheet.cell(source_row, col)
            row_styles.append(
                {
                    "font": copy(source.font),
                    "fill": copy(source.fill),
                    "border": copy(source.border),
                    "alignment": copy(source.alignment),
                    "number_format": source.number_format,
                    "protection": copy(source.protection),
                }
            )
        section_styles.append((sheet.row_dimensions[source_row].height, row_styles))

    clear_product_area(sheet)

    def apply_saved_style(style_index: int, row: int) -> None:
        height, row_styles = section_styles[style_index]
        if height:
            sheet.row_dimensions[row].height = height
        for col, style in enumerate(row_styles, start=1):
            cell = sheet.cell(row, col)
            cell.font = copy(style["font"])
            cell.fill = copy(style["fill"])
            cell.border = copy(style["border"])
            cell.alignment = copy(style["alignment"])
            cell.number_format = style["number_format"]
            cell.protection = copy(style["protection"])

    products_by_category: dict[str, list[MasterProduct]] = {category: [] for category in CATEGORIES}
    extra_categories: list[str] = []
    for product in products:
        category = product.category or "雜貨"
        if category not in products_by_category:
            products_by_category[category] = []
            extra_categories.append(category)
        products_by_category[category].append(product)

    input_rows: list[int] = []
    row = 7
    for category in CATEGORIES + extra_categories:
        category_products = sorted(products_by_category.get(category, []), key=lambda item: item.name)
        if not category_products:
            continue

        apply_saved_style(0, row)
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        sheet.cell(row, 1, category)
        row += 1

        apply_saved_style(1, row)
        for col, header in enumerate(PRODUCT_TABLE_HEADERS, start=1):
            sheet.cell(row, col, header)
        row += 1

        for product in category_products:
            apply_saved_style(2, row)
            sheet.cell(row, 1, None)
            sheet.cell(row, 2, None)
            sheet.cell(row, 3, product.name)
            sheet.cell(row, 4, product.category)
            sheet.cell(row, 5, product.part or product.name)
            sheet.cell(row, 6, None)
            sheet.cell(row, 7, None)
            sheet.cell(row, 8, None)
            input_rows.append(row)
            row += 1

        row += 2

    apply_saved_style(3, row)
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    sheet.cell(row, 1, "總結")
    row += 1

    apply_saved_style(4, row)
    sheet.cell(row, 1, "總售出数量")
    sheet.cell(row, 2, "訂單總收入 HKD")
    row += 1

    apply_saved_style(5, row)
    first_input_row = min(input_rows) if input_rows else 9
    last_input_row = max(input_rows) if input_rows else 9
    sheet.cell(row, 1, f"=SUM(A{first_input_row}:A{last_input_row})")
    sheet.cell(row, 2, f"=SUMPRODUCT(IFERROR(A{first_input_row}:A{last_input_row}*(F{first_input_row}:F{last_input_row}+G{first_input_row}:G{last_input_row}+H{first_input_row}:H{last_input_row}),0))")

    return input_rows


def refresh_order_template(template_path: Path = TEMPLATE_PATH, master_path: Path = MASTER_DATA_PATH) -> Path:
    ensure_master_data(master_path, template_path)
    products = [product for product in load_master_products(master_path).values() if product.active]
    customers = sorted(load_master_customers(master_path))

    workbook = load_workbook(template_path)
    sheet = workbook[workbook.sheetnames[0]]
    input_rows = rebuild_product_area(sheet, products)

    if "主資料" in workbook.sheetnames:
        del workbook["主資料"]
    data_sheet = workbook.create_sheet("主資料")
    data_sheet.sheet_state = "hidden"
    data_sheet.append(["貨品", "分類", "客戶", "訂單狀態"])
    max_rows = max(len(products), len(customers), len(ORDER_STATUSES), 1)
    for index in range(max_rows):
        product = products[index] if index < len(products) else None
        data_sheet.append(
            [
                product.name if product else "",
                product.category if product else "",
                customers[index] if index < len(customers) else "",
                ORDER_STATUSES[index] if index < len(ORDER_STATUSES) else "",
            ]
        )

    sheet.data_validations.dataValidation = []
    product_end = max(2, len(products) + 1)
    category_end = max(2, len(CATEGORIES) + 1)
    customer_end = max(2, len(customers) + 1)
    status_end = max(2, len(ORDER_STATUSES) + 1)

    for index, category in enumerate(CATEGORIES, start=2):
        data_sheet.cell(index, 2, category)

    product_validation = DataValidation(type="list", formula1=f"='主資料'!$A$2:$A${product_end}", allow_blank=True)
    category_validation = DataValidation(type="list", formula1=f"='主資料'!$B$2:$B${category_end}", allow_blank=True)
    status_validation = DataValidation(type="list", formula1=f"='主資料'!$D$2:$D${status_end}", allow_blank=False)
    customer_validation = DataValidation(type="list", formula1=f"='主資料'!$C$2:$C${customer_end}", allow_blank=True)

    sheet.add_data_validation(product_validation)
    sheet.add_data_validation(category_validation)
    sheet.add_data_validation(status_validation)
    sheet.add_data_validation(customer_validation)

    for row in input_rows:
        product_validation.add(sheet.cell(row, 3))
        category_validation.add(sheet.cell(row, 4))
    status_validation.add(sheet["C5"])
    customer_validation.add(sheet["B5"])

    for row in sheet.iter_rows():
        for cell in row:
            cell.protection = Protection(locked=True)

    unlocked_cells = ["B5", "C5", "F5", "H5"]
    for address in unlocked_cells:
        sheet[address].protection = Protection(locked=False)
    for row in input_rows:
        for col in [1, 2, 3, 4, 5, 6, 7, 8]:
            sheet.cell(row, col).protection = Protection(locked=False)

    sheet.protection.sheet = True
    sheet.protection.enable()

    workbook.save(template_path)
    return template_path


def order_template_paths() -> list[Path]:
    seen: set[Path] = set()
    paths: list[Path] = []
    for path in [ORDER_TEMPLATE_PATH, LEGACY_TEMPLATE_PATH, PACKAGE_TEMPLATE_PATH]:
        if path.exists() and path not in seen:
            seen.add(path)
            paths.append(path)
    return paths


def refresh_all_order_templates(master_path: Path = MASTER_DATA_PATH) -> list[Path]:
    paths = order_template_paths()
    if not paths:
        return [refresh_order_template(TEMPLATE_PATH, master_path)]
    return [refresh_order_template(path, master_path) for path in paths]


def main() -> int:
    master_path = ensure_master_data()
    refreshed_paths = refresh_all_order_templates()
    print(f"Master data: {master_path}")
    for path in refreshed_paths:
        print(f"Order template refreshed: {path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
